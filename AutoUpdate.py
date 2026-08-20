import os
from datetime import datetime
from pathlib import Path
import subprocess
import openpyxl
import pandas as pd
from google.oauth2.service_account import Credentials
import gspread

# ==========================================
# CONFIGURATION & PATH MAPPINGS
# ==========================================
BASE_DIR = Path(__file__).resolve().parent  # 'H:\My Drive\Coding Script\Streamlit'

# Exact Google Workspace Shared drive & folder locations
SHARED_DRIVE_OPERATION = Path(r"H:\Shared drives\OPERATION")
LP_FOLDER = SHARED_DRIVE_OPERATION / "LP"
DELIVERY_PLAN_FOLDER = SHARED_DRIVE_OPERATION / "Delivery Plan"
LOTASK_FILE = DELIVERY_PLAN_FOLDER / "LOTask.xlsx"

# Subfolder destination for the local CSV export
OUTPUT_DIR = BASE_DIR / "my_streamlit_app"
CSV_FILENAME = "Extract_Dispatch_Data.csv"
git_relative_path = f"my_streamlit_app/{CSV_FILENAME}"

# Git configuration details
GIT_USER_NAME = "myscholar2298-coder"
GIT_USER_EMAIL = "myscholar2298@gmail.com"

def clean_base_code(name_str):
    s = str(name_str).strip().upper().replace("SAMPLE ", "").replace("SAMPLE_", "")
    return s.split("_")[0].strip()

def run_daily_compilation():
    if not LP_FOLDER.exists():
        print(f"[{datetime.now()}] Error: LP folder not found at {LP_FOLDER}")
        return None

    file_paths = [f for f in LP_FOLDER.glob("*.xls*") if "sales status" in f.name.lower() and "~$" not in f.name]
    if not file_paths:
        print(f"[{datetime.now()}] Warning: No Sales Status files found in LP folder.")
        return None

    master_data = []

    for file_path in file_paths:
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            target_sheets = [sheet for sheet in wb.sheetnames if "^" in sheet]

            for sheet_name in target_sheets:
                ws = wb[sheet_name]
                
                metrics = {}
                for r in range(2, 7):
                    for c in range(1, ws.max_column + 1):
                        cell_val = ws.cell(row=r, column=c).value
                        if cell_val:
                            lbl_str = str(cell_val).strip().upper()
                            next_val = ws.cell(row=r, column=c + 1).value
                            if next_val is not None:
                                try:
                                    num_val = float(next_val)
                                    col_letter = openpyxl.utils.get_column_letter(c)
                                    if col_letter not in metrics:
                                        metrics[col_letter] = {
                                            "SAMPLE BALANCE": None, "SAMPLE IN HAND": None,
                                            "STOCK BALANCE": None, "STOCK IN HAND": None
                                        }
                                    for key in metrics[col_letter].keys():
                                        if key in lbl_str:
                                            metrics[col_letter][key] = num_val
                                except ValueError:
                                    pass

                header_row = 10
                book_pairs = []
                columns_checked = set()

                for col in range(7, ws.max_column + 1):
                    if col in columns_checked:
                        continue
                    h_val = ws.cell(row=header_row, column=col).value
                    if not h_val:
                        continue
                    h_str = str(h_val).strip()
                    h_upper = h_str.upper()

                    if any(x in h_upper for x in ["DELIVERED", "MARKET", "1ST", "2ND", "PROGRESS", "LAIN"]):
                        continue

                    if "_" in h_str or "PA1" in h_upper or "PA3" in h_upper or "SJH3" in h_upper or "ARAB" in h_upper:
                        is_sample = h_upper.startswith("SAMPLE ") or h_upper.startswith("SAMPLE_")
                        clean_title = h_str.replace("SAMPLE ", "").replace("SAMPLE_", "").strip()

                        s_col, q_col = None, None
                        if is_sample:
                            s_col = col
                            columns_checked.add(col)
                            base_code = clean_base_code(h_str)
                            for ahead_col in range(col + 1, col + 12):
                                ahead_val = ws.cell(row=header_row, column=ahead_col).value
                                if ahead_val:
                                    ahead_str = str(ahead_val).strip()
                                    ahead_upper = ahead_str.upper()
                                    if clean_base_code(ahead_str) == base_code and not (ahead_upper.startswith("SAMPLE ") or ahead_upper.startswith("SAMPLE_")):
                                        if not any(x in ahead_upper for x in ["DELIVERED", "MARKET", "1ST", "2ND"]):
                                            q_col = ahead_col
                                            columns_checked.add(ahead_col)
                                            break
                        else:
                            q_col = col
                            columns_checked.add(col)

                        book_pairs.append({
                            "display_name": clean_title,
                            "sample_col": s_col,
                            "qty_col": q_col
                        })

                for row in range(11, ws.max_row + 1):
                    route = ws.cell(row=row, column=1).value
                    school = ws.cell(row=row, column=2).value
                    teacher = ws.cell(row=row, column=3).value

                    if not route and not school:
                        continue

                    for bp in book_pairs:
                        s_idx = bp["sample_col"]
                        q_idx = bp["qty_col"]

                        val_sample = ws.cell(row=row, column=s_idx).value if s_idx else None
                        val_actual = ws.cell(row=row, column=q_idx).value if s_idx else None

                        cell_s_str = str(val_sample).strip() if val_sample is not None else ""
                        cell_q_str = str(val_actual).strip() if val_actual is not None else ""

                        is_zero_sample = (cell_s_str == "" or cell_s_str == "0" or cell_s_str == "0.0" or cell_s_str.lower() == "nan")
                        is_zero_actual = (cell_q_str == "" or cell_q_str == "0" or cell_q_str == "0.0" or cell_q_str.lower() == "nan")

                        is_text_in_sample = False
                        if val_sample is not None and not is_zero_sample:
                            try:
                                float(str(val_sample))
                            except ValueError:
                                if cell_s_str != "#": is_text_in_sample = True

                        is_text_in_qty = False
                        numeric_qty = 0.0
                        is_negative_qty = False
                        if val_actual is not None and not is_zero_actual:
                            try:
                                numeric_qty = float(val_actual)
                                if numeric_qty < 0: is_negative_qty = True
                            except ValueError:
                                if cell_q_str != "#": is_text_in_qty = True

                        has_sample = (not is_zero_sample) or is_text_in_sample
                        has_actual = (not is_zero_actual) or is_text_in_qty or is_negative_qty
                        is_hard_no_stock = (cell_s_str == "#" or cell_q_str == "#" or "NO STOCK" in cell_s_str.upper() or "NO STOCK" in cell_q_str.upper())

                        if has_sample or has_actual or is_hard_no_stock:
                            smp_out = ""
                            smp_text_remark = ""
                            if val_sample is not None and not is_zero_sample:
                                if is_text_in_sample:
                                    smp_text_remark = cell_s_str
                                else:
                                    try:
                                        smp_out = int(float(val_sample)) if float(val_sample).is_integer() else cell_s_str
                                    except ValueError:
                                        smp_text_remark = cell_s_str

                            qty_out = ""
                            qty_text_remark = ""
                            if val_actual is not None and not is_zero_actual:
                                if is_text_in_qty:
                                    qty_text_remark = cell_q_str
                                else:
                                    try:
                                        qty_out = int(numeric_qty) if numeric_qty.is_integer() else cell_q_str
                                    except ValueError:
                                        qty_text_remark = cell_q_str

                            sample_metrics = metrics.get(openpyxl.utils.get_column_letter(s_idx), {}) if s_idx else {}
                            stock_metrics = metrics.get(openpyxl.utils.get_column_letter(q_idx), {}) if q_idx else {}

                            is_sample_no, is_stock_no = False, False
                            is_sample_low, is_stock_low = False, False

                            if has_sample and not is_text_in_sample:
                                if sample_metrics.get("SAMPLE IN HAND") == 0:
                                    is_sample_no = True
                                elif sample_metrics.get("SAMPLE BALANCE") is not None and sample_metrics.get("SAMPLE BALANCE") < 5:
                                    is_sample_low = True

                            if has_actual and not is_negative_qty and not is_text_in_qty:
                                if stock_metrics.get("STOCK IN HAND") == 0:
                                    is_stock_no = True
                                elif stock_metrics.get("STOCK BALANCE") is not None and stock_metrics.get("STOCK BALANCE") < 5:
                                    is_stock_low = True

                            if cell_q_str == "#" or "NO STOCK" in cell_q_str.upper(): is_stock_no = True
                            if cell_s_str == "#" or "NO STOCK" in cell_s_str.upper(): is_sample_no = True

                            computed_remarks = []
                            if is_sample_no and is_stock_no:
                                computed_remarks.append("NO SAMPLE AND NO STOCK")
                            else:
                                if is_sample_no: computed_remarks.append("NO SAMPLE")
                                if is_stock_no: computed_remarks.append("NO STOCK")
                            if is_sample_low: computed_remarks.append("LOW SAMPLE")
                            if is_stock_low: computed_remarks.append("LOW STOCK")

                            if smp_text_remark: computed_remarks.append(smp_text_remark)
                            if qty_text_remark: computed_remarks.append(qty_text_remark)

                            seen = set()
                            unique_rmks = []
                            for rmk in computed_remarks:
                                norm = rmk.strip().upper()
                                if norm and norm not in seen and norm != "NAN":
                                    seen.add(norm)
                                    unique_rmks.append(norm)
                            remark_final = " / ".join(unique_rmks) if unique_rmks else ""

                            master_data.append([
                                route if route is not None else "",
                                school if school is not None else "",
                                teacher if teacher is not None else "",
                                bp["display_name"],
                                smp_out if smp_out is not None else "",
                                qty_out if qty_out is not None else "",
                                remark_final
                            ])
        except Exception as e:
            print(f"[{datetime.now()}] Error reading {file_path.name}: {e}")

    if master_data:
        return pd.DataFrame(master_data, columns=["Route", "School Name", "Teacher", "Book", "Sample", "Qty", "Remark"])
    return None

def run_pipeline_and_sync():
    print(f"\n[{datetime.now()}] Starting compilation and GitHub sync pipeline...")
    
    daily_df = run_daily_compilation()
    if daily_df is None or daily_df.empty:
        print(f"[{datetime.now()}] No records compiled from LP files.")
        return

    lo_task_df = None
    if LOTASK_FILE.exists():
        try:
            lo_task_df = pd.read_excel(LOTASK_FILE)
            lo_task_df.columns = lo_task_df.columns.str.strip()
        except Exception as e:
            print(f"[{datetime.now()}] Error reading LOTask.xlsx: {e}")

    final_processed_rows = []

    for _, row in daily_df.iterrows():
        current_route = str(row.get("Route", "")).strip()
        current_school = str(row.get("School Name", "")).strip()
        current_teacher = str(row.get("Teacher", "")).strip()
        book = str(row.get("Book", "")).strip()
        raw_qty = row.get("Qty", "")
        raw_sample = str(row.get("Sample", "")).strip()
        raw_remark = str(row.get("Remark", "")).strip()
        
        if not current_school or current_school.lower() == "nan":
            continue

        combined_text = f"{raw_remark} {raw_qty} {raw_sample}".upper()

        is_negative_qty = False
        try:
            if raw_qty != "" and float(raw_qty) < 0:
                is_negative_qty = True
        except (ValueError, TypeError):
            pass

        task_types = []
        if "PAYMENT" in combined_text: task_types.append("Payment")
        if "RETURN" in combined_text or is_negative_qty: task_types.append("Return")
        if "CHEQUE" in combined_text: task_types.append("Cheque")

        if not task_types: task_types = ["Delivery"]

        for t_type in task_types:
            q_val, s_val = raw_qty, raw_sample
            if t_type.upper() in str(q_val).upper(): q_val = ""
            if t_type.upper() in str(s_val).upper(): s_val = ""

            clean_remark = raw_remark
            for kw in ["PAYMENT", "RETURN", "CHEQUE"]:
                clean_remark = clean_remark.replace(kw, "").replace(kw.capitalize(), "").strip()
            clean_remark = clean_remark.strip(" /").strip()

            final_processed_rows.append({
                "Date": "",
                "Route": current_route if current_route.lower() != "nan" else "",
                "School Name": current_school if current_school.lower() != "nan" else "",
                "Teacher": current_teacher if current_teacher and current_teacher.lower() != "nan" else "",
                "Task": t_type,
                "Title/Panitia": book if book and book.lower() != "nan" else "",
                "Sample": s_val if s_val != "" and str(s_val).lower() != "nan" else "",
                "Qty": q_val if q_val != "" and str(q_val).lower() != "nan" else "",
                "#Delivery": "",
                "Remark": clean_remark if clean_remark.lower() != "nan" else "",
                "Route_Internal": current_route if current_route.lower() != "nan" else ""
            })

    if lo_task_df is not None:
        for _, row in lo_task_df.iterrows():
            current_school = str(row.get("School Name", row.get("SCHOOL NAME", ""))).strip()
            if not current_school or current_school.lower() == "nan": continue
            
            date_val = row.get("Date", row.get("DATE", ""))
            panitia_val = str(row.get("Panitia", row.get("PANITIA", ""))).strip()
            raw_qty = row.get("Qty", row.get("QTY", ""))
            teacher_val = str(row.get("Teacher", row.get("TEACHER", ""))).strip()
            remark_val = str(row.get("Remark", row.get("REMARK", ""))).strip()
            delivery_val = str(row.get("#Delivery", row.get("#DELIVERY", ""))).strip()
            route_val = str(row.get("Route", row.get("ROUTE", ""))).strip()
            task_val = str(row.get("Task", row.get("TASK", "Delivery"))).strip()

            if pd.notnull(date_val) and isinstance(date_val, datetime):
                date_val = date_val.strftime("%Y-%m-%d")
            else:
                date_val = str(date_val).strip()

            final_processed_rows.append({
                "Date": date_val if date_val.lower() != "nan" else "",
                "Route": route_val if route_val.lower() != "nan" else "",
                "School Name": current_school if current_school.lower() != "nan" else "",
                "Teacher": teacher_val if teacher_val and teacher_val.lower() != "nan" else "",
                "Task": task_val if task_val and task_val.lower() != "nan" else "Delivery",
                "Title/Panitia": panitia_val if panitia_val.lower() != "nan" else "",
                "Sample": "",
                "Qty": raw_qty if raw_qty != "" and str(raw_qty).lower() != "nan" else "",
                "#Delivery": delivery_val if delivery_val.lower() != "nan" else "",
                "Remark": remark_val if remark_val.lower() != "nan" else "",
                "Route_Internal": route_val if route_val.lower() != "nan" else ""
            })

    if not final_processed_rows:
        print(f"[{datetime.now()}] No final rows to export.")
        return

    final_df = pd.DataFrame(final_processed_rows)
    final_df = final_df.replace(["nan", "NaN", "NAN", "None", None, float("nan")], "").fillna("")
    final_df["Route_Internal"] = final_df["Route_Internal"].astype(str).replace("nan", "")
    final_df = final_df.sort_values(by=["Route_Internal", "School Name", "Title/Panitia"])
    final_df["Group"] = final_df["Route_Internal"].str.extract(r'^([A-Za-z])')[0].str.upper().fillna("")

    base_cols = ["Group", "Date", "School Name", "Teacher", "Task", "Title/Panitia", "Sample", "Qty", "#Delivery", "Remark"]
    final_cols = [c for c in base_cols if c in final_df.columns] + ["Route"]
    
    output_df = final_df[final_cols].copy()
    output_df = output_df.replace(["nan", "NaN", "NAN", "None", None, float("nan")], "").fillna("")

    # 1. Save Local CSV copy
    try:
        OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
        local_csv_path = OUTPUT_DIR / CSV_FILENAME
        output_df.to_csv(local_csv_path, index=False, encoding="utf-8-sig", na_rep="")
        print(f"[{datetime.now()}] Successfully saved local CSV copy at: {local_csv_path}")
    except Exception as e:
        print(f"[{datetime.now()}] Error saving local CSV copy: {e}")
        return

    # 2. Sync to Google Sheets
    try:
        SCOPES = ["https://www.googleapis.com/auth/spreadsheets", "https://www.googleapis.com/auth/drive"]
        creds_path = BASE_DIR / "credentials.json"
        creds = Credentials.from_service_account_file(str(creds_path), scopes=SCOPES)
        client = gspread.authorize(creds)
        sheet = client.open("MyScholar_Operations_Live").sheet1
        sheet.clear()
        sheet.append_rows([output_df.columns.values.tolist()] + output_df.values.tolist(), value_input_option='USER_ENTERED')
        print(f"[{datetime.now()}] Successfully updated Google Sheet.")
    except Exception as e:
        if "<Response [200]>" not in str(e):
            print(f"[{datetime.now()}] Google Sheet sync notice: {e}")

    # 3. Automatic Git Commit and Push
    os.chdir(BASE_DIR)
    subprocess.run(f'git config user.name "{GIT_USER_NAME}"', shell=True)
    subprocess.run(f'git config user.email "{GIT_USER_EMAIL}"', shell=True)

    status_res = subprocess.run(f"git status {git_relative_path} --porcelain", capture_output=True, text=True, shell=True)
    if not status_res.stdout.strip():
        print(f"[{datetime.now()}] No content changes found in CSV for Git sync.")
        return

    print(f"[{datetime.now()}] Changes detected. Staging and pushing to GitHub...")
    subprocess.run(f"git add {git_relative_path}", shell=True)
    commit_msg = f"Auto-update dispatch data: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    subprocess.run(f'git commit -m "{commit_msg}"', shell=True)
    push_res = subprocess.run("git push origin main", capture_output=True, text=True, shell=True)
    
    if push_res.returncode == 0:
        print(f"[{datetime.now()}] Successfully synced CSV to GitHub!")
    else:
        print(f"[{datetime.now()}] Error pushing to GitHub: {push_res.stderr}")

if __name__ == "__main__":
    run_pipeline_and_sync()